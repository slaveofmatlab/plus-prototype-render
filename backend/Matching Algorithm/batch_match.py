# -*- coding: utf-8 -*-
"""
功能模块二：批量匹配

- 上传客户询价表单（.xlsx / .xls）
- 自动识别商品名称列，批量匹配标准商品库
- 匹配结果以分页表单展示，每一行的匹配结果为一个下拉框，
  点击可展开其他 Top 召回候选并手动切换，切换后该行编码同步更新
- 导出: 基于上传的原表，在用户选中的商品列后插入"匹配商品品名"和"标准产品编码"两列

可独立运行，也可通过 build_ui() 集成到主应用的标签页中。
"""

import os
import sys
import math
import tempfile
from functools import partial

import pandas as pd
import gradio as gr

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from service import get_matcher

# 复用 Pre-process 模块对查询归一化（提取品牌/规格/核心名，用于置信度评分）
_PREPROCESS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               '..', 'Pre-process', 'Product_Normalizer2.0')
sys.path.insert(0, os.path.abspath(_PREPROCESS_DIR))
from normalizer.main import process_single_record


# ============================================
# 常量
# ============================================

# 商品名称列的识别关键词（按优先级）
_NAME_KEYWORDS_STRONG = [
    '商品名称', '品名', '产品名称', '商品名', '货品名称',
    '物料名称', '品种名称', '货物名称', '物资名称',
]

# 分页: 每页展示的行数（每行一个独立下拉框）
PAGE_SIZE = 15

# 未匹配到候选时下拉框的占位选项（若无任何选项，Gradio 会因 value 不在 choices 内显示红色"错误"徽标）
NO_MATCH_LABEL = "未匹配到结果"


# ============================================
# 表单读取与列识别
# ============================================

def _detect_name_column(df):
    """
    自动识别商品名称列，返回列标签；识别失败返回 None
    """
    # 强关键词优先
    for kw in _NAME_KEYWORDS_STRONG:
        for c in df.columns:
            if kw in str(c):
                return c
    # 宽松回退: 含"名称"或"name"
    for c in df.columns:
        if '名称' in str(c) or 'name' in str(c).lower():
            return c
    return None


def _load_sheet_auto_header(file_path, sheet_name, max_scan=6):
    """
    读取指定 sheet，自动探测表头行（前 max_scan 行内包含名称关键词的行）
    """
    try:
        raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    except Exception:
        return None
    if raw is None or raw.empty:
        return None

    header_row = 0
    for i in range(min(max_scan, len(raw))):
        vals = [str(v) for v in raw.iloc[i].tolist()]
        if any(any(kw in v for kw in _NAME_KEYWORDS_STRONG) for v in vals):
            header_row = i
            break

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
    # 去除完全空行
    df = df.dropna(how='all').reset_index(drop=True)
    return df


def _load_all_sheets(file_path):
    """读取所有非空 sheet，返回 {sheet_name: df}"""
    xl = pd.ExcelFile(file_path)
    sheets = {}
    for sn in xl.sheet_names:
        df = _load_sheet_auto_header(file_path, sn)
        if df is not None and len(df) > 0 and len(df.columns) > 0:
            sheets[sn] = df
    return sheets


def _common_columns(sheets_state, selected_sheets):
    """计算所选工作表的公共列标签（保持首表顺序）；无交集时回退到首表全部列"""
    if not selected_sheets:
        return []
    first_cols = [str(c) for c in sheets_state[selected_sheets[0]].columns]
    if len(selected_sheets) == 1:
        return first_cols
    common_set = set(first_cols)
    for sn in selected_sheets[1:]:
        common_set &= {str(c) for c in sheets_state[sn].columns}
    common = [c for c in first_cols if c in common_set]
    return common if common else first_cols


def _detect_common_name_column(sheets_state, selected_sheets):
    """在所选工作表中识别一个通用的商品名称列标签"""
    if not selected_sheets:
        return None
    common = _common_columns(sheets_state, selected_sheets)
    first_col = _detect_name_column(sheets_state[selected_sheets[0]])
    if first_col is not None and str(first_col) in common:
        return str(first_col)
    for sn in selected_sheets:
        c = _detect_name_column(sheets_state[sn])
        if c is not None and str(c) in common:
            return str(c)
    return str(first_col) if first_col is not None else None


def _get_column_by_label(df, label):
    """根据列标签字符串找到实际列对象"""
    for c in df.columns:
        if str(c) == label:
            return c
    return None


def _get_name_series(df, col):
    """获取名称列 Series（兼容重复列名）"""
    s = df[col]
    if isinstance(s, pd.DataFrame):  # 重复列名时返回 DataFrame
        s = s.iloc[:, 0]
    return s


# ============================================
# 上传处理
# ============================================

def on_file_upload(file):
    """
    文件上传后: 读取所有 sheet，自动选择最佳 sheet 与名称列
    返回: sheets_state, sheet下拉, 列下拉, 预览, 状态
    """
    if file is None:
        return ({}, gr.update(choices=[], value=None),
                gr.update(choices=[], value=None), None, "请先上传表单文件")

    file_path = file.name if hasattr(file, 'name') else str(file)

    try:
        sheets = _load_all_sheets(file_path)
    except Exception as e:
        return ({}, gr.update(choices=[], value=None),
                gr.update(choices=[], value=None), None, f"读取文件失败: {e}")

    if not sheets:
        return ({}, gr.update(choices=[], value=None),
                gr.update(choices=[], value=None), None, "未读取到有效数据表")

    # 默认选中所有能识别出名称列的工作表（便于多子表合并处理），用户可手动增减
    valid_sheets = [sn for sn, df in sheets.items() if _detect_name_column(df) is not None]
    default_selected = valid_sheets if valid_sheets else list(sheets.keys())

    col_choices = _common_columns(sheets, default_selected)
    best_col = _detect_common_name_column(sheets, default_selected)
    preview = sheets[default_selected[0]].head(10) if default_selected else None
    total_rows = sum(len(sheets[s]) for s in default_selected)

    status = (f"已加载 {len(sheets)} 个工作表 | 已选 {len(default_selected)} 个 | "
              f"商品名称列: {best_col} | 合计 {total_rows} 行")

    return (sheets,
            gr.update(choices=list(sheets.keys()), value=default_selected),
            gr.update(choices=col_choices, value=best_col),
            preview, status)


def on_sheet_change(selected_sheets, sheets_state):
    """切换所选工作表后: 更新列下拉与预览（支持多选合并处理）"""
    if isinstance(selected_sheets, str):
        selected_sheets = [selected_sheets]
    if not sheets_state or not selected_sheets:
        return gr.update(choices=[], value=None), None, "请至少选择一个工作表"

    selected_sheets = [s for s in selected_sheets if s in sheets_state]
    if not selected_sheets:
        return gr.update(choices=[], value=None), None, "无效的工作表"

    col_choices = _common_columns(sheets_state, selected_sheets)
    best_col = _detect_common_name_column(sheets_state, selected_sheets)
    preview = sheets_state[selected_sheets[0]].head(10)
    total_rows = sum(len(sheets_state[s]) for s in selected_sheets)
    status = (f"已选 {len(selected_sheets)} 个工作表: {', '.join(selected_sheets)} | "
              f"商品名称列: {best_col} | 合计 {total_rows} 行")

    return gr.update(choices=col_choices, value=best_col), preview, status


# ============================================
# 批量匹配
# ============================================

def _candidate_label(cand):
    """生成候选下拉选项的显示文本（得分/置信度已拆分到独立列，此处仅展示排名与品名）"""
    return f"[{cand['rank']}] {cand['标准产品名称']}"


def _fmt_confidence(v):
    """置信度展示格式（空值显示 '-'）"""
    try:
        return f"{float(v):.2f}"
    except (TypeError, ValueError):
        return '-'


def _fmt_recall(v):
    """召回得分展示格式（空值显示 '-'）"""
    try:
        return f"{float(v):.4f}"
    except (TypeError, ValueError):
        return '-'


def _current_label(row):
    """获取某行当前选中候选对应的下拉标签"""
    for c in row['candidates']:
        if (c['标准产品名称'] == row['selected_name']
                and str(c['标准产品编码']) == row['selected_code']):
            return c['_label']
    return None


def _parse_score(row):
    """解析行当前选中候选的置信度（优先），无置信度时回退召回得分；无法解析返回 None"""
    for key in ('selected_confidence', 'selected_score'):
        try:
            return float(row[key])
        except (TypeError, ValueError, KeyError):
            continue
    return None


def _page_updates(state, page, filter_status=None):
    """
    生成某一页所有行槽位组件的更新 + 分页信息 + 页码状态
    每行 5 个输出: [label, dd, code, recall, conf]
    返回顺序: 各行槽位依次 + page_info + page + page_input
    filter_status: 非 None 时追加到末尾，供筛选事件使用
    """
    total = len(state)
    total_pages = max(1, math.ceil(total / PAGE_SIZE)) if total > 0 else 1
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    # 多工作表合并时，在行标签前缀标注来源子表，便于区分
    show_sheet = len({r.get('sheet') for r in state}) > 1

    updates = []
    for i in range(PAGE_SIZE):
        gidx = start + i
        if gidx < total:
            r = state[gidx]
            labels = [c['_label'] for c in r['candidates']]
            if show_sheet:
                updates.append(gr.update(visible=True, value=f"{r['seq']}. [{r.get('sheet')}] {r['customer_name']}"))
            else:
                updates.append(gr.update(visible=True, value=f"{r['seq']}. {r['customer_name']}"))
            if labels:
                updates.append(gr.update(visible=True, choices=labels, value=_current_label(r)))
                updates.append(gr.update(visible=True, value=str(r['selected_code'])))
                updates.append(gr.update(visible=True, value=_fmt_recall(r.get('selected_score'))))
                updates.append(gr.update(visible=True, value=_fmt_confidence(r.get('selected_confidence'))))
            else:
                # 未匹配到任何候选：下拉显示占位，其余列隐藏
                updates.append(gr.update(visible=True, choices=[NO_MATCH_LABEL], value=NO_MATCH_LABEL))
                updates.append(gr.update(visible=False))
                updates.append(gr.update(visible=False))
                updates.append(gr.update(visible=False))
        else:
            # 空槽位: 隐藏整行组件（避免空 choices 报错），并提供合法占位值
            updates.append(gr.update(visible=False, value=""))
            updates.append(gr.update(visible=False, choices=[''], value=''))
            updates.append(gr.update(visible=False, value=""))
            updates.append(gr.update(visible=False, value=""))
            updates.append(gr.update(visible=False, value=""))

    page_info = (f"第 {page + 1} / {total_pages} 页（共 {total} 条）"
                 if total > 0 else "暂无数据")
    updates.append(page_info)
    updates.append(page)  # page_state
    updates.append(page + 1)  # page_input（1-based 页码）
    if filter_status is not None:
        updates.append(filter_status)
    return updates


def _empty_result(status):
    """批量匹配出错/无数据时的统一返回（空 state + 空页面 + 重置筛选 + 状态）"""
    return tuple([[]] + _page_updates([], 0) + [None, None, "", status])


def run_batch_match(sheets_state, selected_sheets, name_column, top_k, progress=gr.Progress()):
    """
    执行批量匹配（支持多工作表合并处理）
    返回: state, 各页行槽位更新, 分页信息, 页码状态, 状态文本
    """
    if isinstance(selected_sheets, str):
        selected_sheets = [selected_sheets]
    if not sheets_state or not selected_sheets:
        return _empty_result("请先上传表单并选择工作表")

    selected_sheets = [s for s in selected_sheets if s in sheets_state]
    if not selected_sheets:
        return _empty_result("无效的工作表")

    # 汇总所有所选工作表的非空名称行: (seq, name, sheet, name_col, orig_idx)
    pending = []
    skipped_sheets = []
    for sn in selected_sheets:
        df = sheets_state[sn]
        col = _get_column_by_label(df, name_column) or _detect_name_column(df)
        if col is None:
            skipped_sheets.append(sn)
            continue
        name_series = _get_name_series(df, col)
        for idx, val in enumerate(name_series.tolist()):
            if pd.isna(val) or not str(val).strip():
                continue
            pending.append((len(pending) + 1, str(val).strip(), sn, str(col), idx))

    if not pending:
        msg = "所选工作表没有有效的商品名称数据"
        if skipped_sheets:
            msg += f"（未识别名称列: {', '.join(skipped_sheets)}）"
        return _empty_result(msg)

    matcher = get_matcher()
    top_k = int(top_k)

    state = []
    total = len(pending)
    for i, (seq, name, sn, name_col, orig_idx) in enumerate(pending):
        progress((i + 1) / total, desc=f"正在匹配 {i + 1}/{total}: {name}")
        # 对查询归一化一次: 供置信度评分使用，并传入匹配器避免重复计算
        query_info = process_single_record(name, brands_sorted=matcher.brands)
        candidates = matcher.query_extended(name, top_n=top_k, query_info=query_info)
        for c in candidates:
            # 置信度（独立于召回重排得分）: 品牌/规格/核心名三维度
            c['confidence'] = matcher.compute_confidence(query_info, c)
            c['_label'] = _candidate_label(c)

        if candidates:
            top1 = candidates[0]
            selected_name = top1['标准产品名称']
            selected_code = str(top1['标准产品编码'])
            selected_score = top1['score']
            selected_confidence = top1['confidence']
        else:
            selected_name = '未匹配'
            selected_code = ''
            selected_score = ''
            selected_confidence = ''

        state.append({
            'seq': seq,
            'customer_name': name,
            'sheet': sn,
            'name_col': name_col,
            'orig_idx': orig_idx,
            'candidates': candidates,
            'selected_name': selected_name,
            'selected_code': selected_code,
            'selected_score': selected_score,
            'selected_confidence': selected_confidence,
            'query_brand': query_info.get('detected_brand') or '',
            'query_spec': query_info.get('normalized_spec') or '',
        })

    matched = sum(1 for r in state if r['selected_name'] != '未匹配')
    status = f"批量匹配完成: 共 {total} 条（{len(selected_sheets)} 个工作表），已匹配 {matched} 条"
    if skipped_sheets:
        status += f" | 跳过未识别名称列的工作表: {', '.join(skipped_sheets)}"
    # 末尾追加: 重置筛选视图与筛选状态
    return tuple([state] + _page_updates(state, 0) + [None, None, "", status])


# ============================================
# 分页与行内切换
# ============================================

def change_page(page_state, match_state, filter_view, delta):
    """翻页: 返回目标页的行槽位更新（delta 作为关键字参数由 partial 绑定，须置于末尾）"""
    new_page = (page_state or 0) + delta
    data = filter_view if filter_view is not None else (match_state or [])
    return _page_updates(data, new_page)


def jump_to_page(target_page, match_state, filter_view):
    """跳转到指定页码（用户输入为 1-based），自动夹紧到有效范围"""
    try:
        target = int(target_page)
    except (TypeError, ValueError):
        target = 1
    data = filter_view if filter_view is not None else (match_state or [])
    return _page_updates(data, target - 1)


def apply_filter(match_state, low, high):
    """
    筛选当前选中候选的置信度（无置信度时回退召回得分）落在区间 [low, high] 内的结果。
    未匹配（无得分）的行在下限 <= 0 时一并纳入。
    返回: [页槽位更新, 分页信息, 页码, filter_status, filter_view, filter_map]
    """
    if not match_state:
        return tuple(_page_updates([], 0) + ["请先执行批量匹配", None, None])
    try:
        lo = float(low)
        hi = float(high)
    except (TypeError, ValueError):
        return tuple(_page_updates(match_state, 0) + ["阈值无效，请输入数字（如 0.3 ~ 0.6）", None, None])
    if lo > hi:
        return tuple(_page_updates(match_state, 0) + ["下限不能大于上限，请调整区间", None, None])

    view, view_map = [], []
    for i, r in enumerate(match_state):
        s = _parse_score(r)
        if s is None:
            # 未匹配/无得分: 下限 <= 0 时纳入（保持“含未匹配”的既有行为）
            if lo <= 0:
                view.append(r)
                view_map.append(i)
        elif lo <= s <= hi:
            view.append(r)
            view_map.append(i)

    filter_status = f"筛选视图: {len(view)} 条置信度在 [{lo:g}, {hi:g}] 区间（共 {len(match_state)} 条）"
    return tuple(_page_updates(view, 0, filter_status) + [view, view_map])


def clear_filter(match_state):
    """清除筛选视图，恢复展示全部结果"""
    total = len(match_state or [])
    fs = f"已清除筛选，显示全部 {total} 条" if total else "暂无数据"
    # 末尾两个 None: 重置 filter_view / filter_map 状态
    return tuple(_page_updates(match_state or [], 0, fs) + [None, None])


def on_row_change(new_value, page_state, match_state, filter_view, filter_map, slot_idx):
    """
    某一行下拉框选择了新的候选: 更新该行匹配结果
    返回: 该行编码框更新, match_state
    """
    if not match_state or not new_value:
        return gr.update(), gr.update(), gr.update(), match_state

    vidx = (page_state or 0) * PAGE_SIZE + slot_idx
    if filter_view is not None:
        # 筛选视图: 通过映射定位 match_state 中的原始行
        if not filter_map or vidx < 0 or vidx >= len(filter_map):
            return gr.update(), gr.update(), gr.update(), match_state
        gidx = filter_map[vidx]
        if gidx < 0 or gidx >= len(match_state):
            return gr.update(), gr.update(), gr.update(), match_state
        row = match_state[gidx]
    else:
        gidx = vidx
        if gidx < 0 or gidx >= len(match_state):
            return gr.update(), gr.update(), gr.update(), match_state
        row = match_state[gidx]
    chosen = None
    for c in row['candidates']:
        if c.get('_label') == new_value:
            chosen = c
            break
    if chosen is None:
        return gr.update(), gr.update(), gr.update(), match_state

    row['selected_name'] = chosen['标准产品名称']
    row['selected_code'] = str(chosen['标准产品编码'])
    row['selected_score'] = chosen['score']
    row['selected_confidence'] = chosen.get('confidence')

    return (gr.update(value=str(row['selected_code'])),
            gr.update(value=_fmt_recall(row.get('selected_score'))),
            gr.update(value=_fmt_confidence(row.get('selected_confidence'))),
            match_state)


# ============================================
# 导出结果（基于原表）
# ============================================

def export_results(match_state, sheets_state, selected_sheets, name_column):
    """
    基于上传的原表导出（支持多工作表）:
    在每个所选工作表的商品名称列后插入"标准商品名称"、"标准产品编码"、"匹配置信度"三列，
    三列用浅蓝底色区分，所有工作表写入同一个 xlsx 文件，返回文件路径。
    """
    from openpyxl.styles import PatternFill

    print(f"[export] match_state: {len(match_state) if match_state else 0} rows")
    print(f"[export] sheets_state keys: {list(sheets_state.keys()) if sheets_state else 'None'}")
    print(f"[export] selected_sheets: {selected_sheets}")
    print(f"[export] name_column: {name_column}")

    if not match_state or not sheets_state:
        print("[export] ERROR: match_state or sheets_state is empty")
        return None
    if isinstance(selected_sheets, str):
        selected_sheets = [selected_sheets]
    if not selected_sheets:
        selected_sheets = list({r['sheet'] for r in match_state if r.get('sheet')})
    print(f"[export] final selected_sheets: {selected_sheets}")

    # 按工作表分组匹配结果
    by_sheet = {}
    for r in match_state:
        sn = r.get('sheet')
        if sn is not None:
            by_sheet.setdefault(sn, []).append(r)

    out_path = os.path.join(tempfile.gettempdir(), '批量匹配结果.xlsx')
    wrote_any = False
    # 浅蓝底色
    blue_fill = PatternFill(start_color='DAEEF8', end_color='DAEEF8', fill_type='solid')

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        for sn in selected_sheets:
            if sn not in sheets_state:
                continue
            df = sheets_state[sn].copy()
            rows = by_sheet.get(sn, [])
            # 定位名称列: 优先用该表实际匹配列，回退到传入列
            name_col = rows[0]['name_col'] if rows and rows[0].get('name_col') else name_column
            col = _get_column_by_label(df, name_col) or _get_column_by_label(df, name_column)
            if col is None:
                df.to_excel(writer, sheet_name=sn[:31], index=False)
                wrote_any = True
                continue
            col_pos = list(df.columns).index(col)
            name_vals = [''] * len(df)
            code_vals = [''] * len(df)
            conf_vals = [''] * len(df)
            for r in rows:
                idx = r.get('orig_idx')
                if idx is not None and 0 <= idx < len(df):
                    name_vals[idx] = r['selected_name']
                    code_vals[idx] = r['selected_code']
                    conf_vals[idx] = _fmt_confidence(r.get('selected_confidence'))
            df.insert(col_pos + 1, '标准商品名称', name_vals)
            df.insert(col_pos + 2, '标准产品编码', code_vals)
            df.insert(col_pos + 3, '匹配置信度', conf_vals)
            df.to_excel(writer, sheet_name=sn[:31], index=False)
            wrote_any = True

            # 为新增的三列设置浅蓝底色
            ws = writer.sheets[sn[:31]]
            header_row = 1  # 1-based, 第一行是表头
            new_col_start = col_pos + 2  # 1-based: col_pos(0-based) + 1(插入第1列) + 1(1-based)
            for c in range(new_col_start, new_col_start + 3):
                # 表头
                ws.cell(row=header_row, column=c).fill = blue_fill
                # 数据行
                for data_row in range(2, len(df) + 2):
                    ws.cell(row=data_row, column=c).fill = blue_fill

    return out_path if wrote_any else None


# ============================================
# 界面构建
# ============================================

def build_ui():
    """在当前 Blocks 上下文中构建批量匹配界面"""
    gr.Markdown("### 批量匹配")
    gr.Markdown("上传客户询价表单（支持含多个子表/Sheet 的文件），自动识别商品名称列并批量匹配标准商品库。"
                "**可在工作表下拉框中多选需要处理的子表进行合并匹配**，匹配结果按页展示，"
                "**每一行的匹配结果都是一个下拉框**，点击可展开其他 Top 候选并手动切换。")

    # 匹配状态（核心数据）
    match_state = gr.State([])
    sheets_state = gr.State({})
    page_state = gr.State(0)  # 当前页码（0 起）
    # 筛选视图状态: 非 None 时页面展示的是筛选后的子集
    filter_view_state = gr.State(None)   # 筛选后的行列表（引用 match_state 中的行）
    filter_map_state = gr.State(None)    # 筛选视图索引 → match_state 索引的映射

    # ---- 上传与参数区 ----
    with gr.Row():
        file_input = gr.File(label="上传询价表单 (.xlsx / .xls)", file_types=['.xlsx', '.xls'], scale=2)
        sheet_dropdown = gr.Dropdown(label="工作表（可多选合并处理）", choices=[], interactive=True, multiselect=True, scale=1)
        column_dropdown = gr.Dropdown(label="商品名称列", choices=[], interactive=True, scale=1)
        topk_slider = gr.Slider(label="备选数量(Top K)", minimum=5, maximum=50, step=5, value=20, scale=1)

    with gr.Row():
        match_btn = gr.Button("开始批量匹配", variant="primary")
        status_text = gr.Textbox(label="状态", interactive=False, scale=4)

    # ---- 预览区 ----
    with gr.Accordion("表单预览（前10行）", open=False):
        preview_df = gr.Dataframe(label="原始表单预览", interactive=False)

    # ---- 匹配结果区（分页 + 每行独立下拉框） ----
    gr.Markdown("#### 匹配结果")
    gr.Markdown("💡 每一行的匹配结果为下拉框，**点击可展开其他 Top 候选并切换**，切换后右侧的编码、召回得分、置信度自动更新。"
                "**召回得分**反映名称/规格/品牌的综合召回相似度；**置信度**（0~1）独立评估匹配结果的正确概率，分数越高越可信。"
                "可输入置信度区间（下限 ~ 上限）**筛选需人工核查的结果**（下限为 0 时未匹配的行也会纳入）。")

    # 置信度区间筛选工具栏
    with gr.Row(elem_id='filter-toolbar'):
        filter_low = gr.Number(label="置信度下限", value=0.0, precision=2, scale=1)
        filter_high = gr.Number(label="置信度上限", value=0.6, precision=2, scale=1)
        filter_btn = gr.Button("筛选区间结果", variant="secondary", scale=1)
        clear_filter_btn = gr.Button("清除筛选", scale=1)
        filter_status_text = gr.Textbox(interactive=False, scale=3, show_label=False)

    # 表头（使用 HTML 确保与表体对齐）
    gr.HTML("""
    <div style="display: flex; gap: 8px; margin-bottom: 8px;">
        <div style="flex: 2; font-weight: 600; padding: 8px 10px; background: #f1f5f9; border-radius: 6px; font-size: 13.5px;">序号 / 客户商品名称</div>
        <div style="flex: 3; font-weight: 600; padding: 8px 10px; background: #f1f5f9; border-radius: 6px; font-size: 13.5px;">匹配商品品名（点击下拉切换备选）</div>
        <div style="flex: 2; font-weight: 600; padding: 8px 10px; background: #f1f5f9; border-radius: 6px; font-size: 13.5px;">标准产品编码</div>
        <div style="flex: 1; font-weight: 600; padding: 8px 10px; background: #f1f5f9; border-radius: 6px; font-size: 13.5px;">召回得分</div>
        <div style="flex: 1; font-weight: 600; padding: 8px 10px; background: #f1f5f9; border-radius: 6px; font-size: 13.5px;">置信度</div>
    </div>
    """)

    # 行槽位（每页 PAGE_SIZE 行）
    row_labels, row_dropdowns, row_codes, row_recalls, row_confs = [], [], [], [], []
    for i in range(PAGE_SIZE):
        with gr.Row(elem_id=f'result-row-{i}'):
            lbl = gr.Textbox(interactive=False, scale=2, show_label=False, elem_id=f'cell-name-{i}')
            dd = gr.Dropdown(choices=[''], interactive=True, scale=3, show_label=False)
            code = gr.Textbox(interactive=False, scale=2, show_label=False, elem_id=f'cell-code-{i}')
            recall = gr.Textbox(interactive=False, scale=1, show_label=False, elem_id=f'cell-recall-{i}')
            conf = gr.Textbox(interactive=False, scale=1, show_label=False, elem_id=f'cell-conf-{i}')
        row_labels.append(lbl)
        row_dropdowns.append(dd)
        row_codes.append(code)
        row_recalls.append(recall)
        row_confs.append(conf)

    # 分页导航
    with gr.Row():
        prev_btn = gr.Button("上一页", scale=1)
        page_info = gr.Textbox(value="暂无数据", interactive=False, scale=2, show_label=False)
        next_btn = gr.Button("下一页", scale=1)
        page_input = gr.Number(value=1, precision=0, scale=1, show_label=False)
        jump_btn = gr.Button("跳转", scale=1)

    # 页槽位组件的输出顺序（与 _page_updates 返回顺序一致，每行 5 个组件）
    page_outputs = []
    for i in range(PAGE_SIZE):
        page_outputs.extend([row_labels[i], row_dropdowns[i], row_codes[i], row_recalls[i], row_confs[i]])
    page_outputs.extend([page_info, page_state, page_input])

    # ---- 导出区 ----
    with gr.Row():
        export_btn = gr.Button("导出匹配结果（基于原表）", variant="secondary")
        export_file = gr.File(label="下载结果文件", interactive=False)

    # ---- 事件绑定 ----
    # 上传文件 → 读取并自动识别
    file_input.upload(
        fn=on_file_upload,
        inputs=[file_input],
        outputs=[sheets_state, sheet_dropdown, column_dropdown, preview_df, status_text],
    )

    # 切换工作表 → 更新列下拉与预览
    sheet_dropdown.change(
        fn=on_sheet_change,
        inputs=[sheet_dropdown, sheets_state],
        outputs=[column_dropdown, preview_df, status_text],
    )

    # 开始批量匹配 → 填充第 0 页（同时重置筛选视图）
    match_btn.click(
        fn=run_batch_match,
        inputs=[sheets_state, sheet_dropdown, column_dropdown, topk_slider],
        outputs=[match_state] + page_outputs + [filter_view_state, filter_map_state, filter_status_text, status_text],
    )

    # 得分区间筛选 / 清除筛选
    filter_btn.click(
        fn=apply_filter,
        inputs=[match_state, filter_low, filter_high],
        outputs=page_outputs + [filter_status_text, filter_view_state, filter_map_state],
    )
    clear_filter_btn.click(
        fn=clear_filter,
        inputs=[match_state],
        outputs=page_outputs + [filter_status_text, filter_view_state, filter_map_state],
    )

    # 翻页（筛选状态下仅在筛选视图内翻页）
    prev_btn.click(
        fn=partial(change_page, delta=-1),
        inputs=[page_state, match_state, filter_view_state],
        outputs=page_outputs,
    )
    next_btn.click(
        fn=partial(change_page, delta=1),
        inputs=[page_state, match_state, filter_view_state],
        outputs=page_outputs,
    )

    # 输入页码跳转
    jump_btn.click(
        fn=jump_to_page,
        inputs=[page_input, match_state, filter_view_state],
        outputs=page_outputs,
    )

    # 每行下拉框切换候选 → 更新该行编码/召回得分/置信度与 state（筛选状态下同步写回原始结果）
    for i, dd in enumerate(row_dropdowns):
        dd.change(
            fn=partial(on_row_change, slot_idx=i),
            inputs=[dd, page_state, match_state, filter_view_state, filter_map_state],
            outputs=[row_codes[i], row_recalls[i], row_confs[i], match_state],
        )

    # 导出结果（基于原表）
    export_btn.click(
        fn=export_results,
        inputs=[match_state, sheets_state, sheet_dropdown, column_dropdown],
        outputs=[export_file],
    )


# ============================================
# 独立运行入口
# ============================================

if __name__ == '__main__':
    print("正在初始化匹配器...")
    get_matcher()
    print("初始化完成，启动批量匹配界面...")

    with gr.Blocks(title="商品匹配 - 批量匹配") as app:
        gr.Markdown("# 商品价格推荐 - 批量匹配")
        build_ui()

    app.launch(server_name="127.0.0.1", server_port=7862, theme=gr.themes.Soft())
