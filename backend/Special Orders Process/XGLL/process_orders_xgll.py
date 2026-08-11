# -*- coding: utf-8 -*-
"""
处理报价单中的商品描述，提取中文商品名称。
转换规则：
1. 去除英文部分
2. 去除类目信息（如：小食、干果、酱料等）
3. 繁体字转简体字
4. 去除储存类型（干货、冰鲜、急冻）和国家信息
5. 去除 "PER PKT"、"PER BTL" 等包装描述
"""

import sys
sys.stdout.reconfigure(encoding='utf-8')
import re
import openpyxl
from opencc import OpenCC

# 繁体转简体转换器
cc = OpenCC('t2s')

# ============ 配置列表 ============

# 储存类型
STORAGE_TYPES = ['干货', '冰鲜', '急冻', '常温']

# 国家/地区列表
COUNTRIES = [
    '中国', '南韩', '韩国', '泰国', '法国', '美国', '印度', '意大利',
    '澳大利亚', '波兰', '白俄罗斯', '日本', '德国', '西班牙', '葡萄牙',
    '荷兰', '比利时', '英国', '加拿大', '墨西哥', '巴西', '阿根廷',
    '新西兰', '越南', '马来西亚', '印度尼西亚', '印尼', '菲律宾',
    '土耳其', '希腊', '埃及', '南非', '智利', '秘鲁', '哥伦比亚',
    '瑞士', '奥地利', '丹麦', '挪威', '瑞典', '芬兰', '爱尔兰',
    '以色列', '沙特', '阿联酋', '新加坡', '缅甸', '柬埔寨',
    '斯里兰卡', '巴基斯坦', '孟加拉', '尼泊尔', '摩洛哥',
    '突尼斯', '阿尔及利亚', '匈牙利', '捷克', '罗马尼亚',
    '保加利亚', '克罗地亚', '塞尔维亚', '乌克兰', '俄罗斯',
    '格鲁吉亚', '亚美尼亚', '阿塞拜疆', '乌兹别克斯坦',
    '哈萨克斯坦', '蒙古', '朝鲜', '老挝', '文莱', '东帝汶',
    '巴布亚新几内亚', '斐济', '萨摩亚', '汤加', '瓦努阿图',
    '哥斯达黎加', '巴拿马', '古巴', '牙买加', '多米尼加',
    '委内瑞拉', '厄瓜多尔', '玻利维亚', '巴拉圭', '乌拉圭',
    '苏里南', '圭亚那', '特立尼达和多巴哥', '巴巴多斯',
    '巴哈马', '伯利兹', '洪都拉斯', '尼加拉瓜', '萨尔瓦多',
    '危地马拉', '海地', '约旦', '黎巴嫩', '叙利亚', '伊拉克',
    '伊朗', '阿富汗', '也门', '阿曼', '巴林', '卡塔尔',
    '科威特', '塞浦路斯', '马耳他', '卢森堡', '摩纳哥',
    '安道尔', '列支敦士登', '圣马力诺', '梵蒂冈', '冰岛',
    '立陶宛', '拉脱维亚', '爱沙尼亚', '斯洛伐克', '斯洛文尼亚',
    '北马其顿', '阿尔巴尼亚', '黑山', '波黑', '科索沃',
    '摩尔多瓦', '塔吉克斯坦', '吉尔吉斯斯坦', '土库曼斯坦',
    '马尔代夫', '不丹', '毛里求斯', '塞舌尔', '马达加斯加',
    '肯尼亚', '坦桑尼亚', '乌干达', '卢旺达', '布隆迪',
    '刚果', '加蓬', '喀麦隆', '尼日利亚', '加纳', '塞内加尔',
    '科特迪瓦', '马里', '布基纳法索', '尼日尔', '乍得',
    '苏丹', '埃塞俄比亚', '索马里', '莫桑比克', '津巴布韦',
    '赞比亚', '马拉维', '安哥拉', '纳米比亚', '博茨瓦纳',
    '莱索托', '斯威士兰', '毛里塔尼亚', '冈比亚', '几内亚',
    '塞拉利昂', '利比里亚', '多哥', '贝宁', '赤道几内亚',
    '圣多美和普林西比', '佛得角', '科摩罗', '吉布提',
    '厄立特里亚', '利比亚', '巴勒斯坦', '库克群岛', '纽埃',
    '帕劳', '马绍尔群岛', '密克罗尼西亚', '基里巴斯', '瑙鲁',
    '图瓦卢', '所罗门群岛', '托克劳', '美属萨摩亚', '关岛',
    '北马里亚纳群岛', '波多黎各', '美属维尔京群岛',
    '英属维尔京群岛', '开曼群岛', '百慕大', '格陵兰',
    '法罗群岛', '直布罗陀', '泽西岛', '根西岛', '马恩岛',
    '奥兰群岛', '斯瓦尔巴群岛', '扬马延岛', '皮特凯恩群岛',
    '圣赫勒拿', '阿森松岛', '特里斯坦-达库尼亚',
    '福克兰群岛', '南乔治亚和南桑威奇群岛',
    '法属波利尼西亚', '新喀里多尼亚', '瓦利斯和富图纳',
    '马约特', '留尼汪', '瓜德罗普', '马提尼克', '法属圭亚那',
    '圣皮埃尔和密克隆', '圣巴泰勒米', '圣马丁',
    '库拉索', '阿鲁巴', '荷属圣马丁', '博内尔',
    '萨巴', '圣尤斯特歇斯', '安圭拉', '蒙特塞拉特',
    '特克斯和凯科斯群岛', '南苏丹', '中非',
]

# 类目列表（从数据中提取的所有类目）
CATEGORIES = [
    '中国白酒', '中式干货', '中药材', '乳制品', '干果', '健康食品',
    '全麦粉', '其他奶制品', '其他面粉', '冰冻蔬菜', '冰淇淋', '冲剂饮料',
    '加工牛肉类', '加工猪肉类', '加工肉类', '加工鸡肉类', '加工鸭肉类',
    '加工鹅肉类', '印尼食品', '厨酒', '可可粉', '吉士粉',
    '味美思白加香葡萄酒', '咖啡', '咖啡粉', '坚果', '处理海鲜',
    '处理过的蔬菜', '奶巧克力', '奶油', '小食', '巧克力', '巧克力壳',
    '干果仁糖', '干菇', '干蔬菜', '干贝', '干香草', '干鱼肚',
    '急冻水果', '意粉', '挞壳', '日式食品', '日本清酒', '日本食品',
    '杂类', '杂货', '松子', '果仁类', '果仁类酱', '果仁酱', '果汁',
    '果酒', '果酱', '植物饮料', '油', '泰国食品', '添加剂', '火腿',
    '点心', '点心材料', '牛奶', '牛奶巧克力', '瓜子仁', '生粉',
    '白兰地', '白巧克力', '谷物类', '竹笙', '米', '籽', '糖', '糖果',
    '糖浆', '糯米粉', '素食品', '罐头水果', '罐头海鲜', '罐头肉',
    '罐头蔬菜', '罐装蔬菜', '肠', '芝士', '芝士酱', '芝麻', '蔬菜',
    '薯条', '薯片', '蛋', '蛋糕装饰', '蜂蜜', '蜜糖', '西米',
    '调味品', '调味料', '豆类', '酥皮', '酱料', '酸奶', '醋',
    '面包', '面包类', '面包粉', '面团', '面类', '面粉', '预拌粉',
    '颜料', '饮料', '饼干', '饼房产品', '香精', '香肠', '马来食品',
    '鸡', '黄油', '黑巧克力', '龟苓膏粉',
    # 繁体类目（转换后可能仍需要的）
    '穀物類', '蛋糕裝飾', '面類', '果醬',
]

# 子类目（跟在主类目后面的修饰词，不是商品名称的一部分）
SUB_CATEGORIES = [
    '整粒', '果茸', '果肉', '果泥', '腌制食品',
]

# 按长度降序排列类目，确保最长匹配优先
CATEGORIES_SORTED = sorted(CATEGORIES, key=len, reverse=True)


def remove_trailing_parenthetical(text):
    """
    移除字符串末尾的括号内容。
    处理嵌套括号的情况，如：
    "小食 香脆椒 (黄飞红) 308GM 干货 中国 (小食 香脆椒 (黄飞红) 308GM 干货 中国)"
    -> "小食 香脆椒 (黄飞红) 308GM 干货 中国"
    """
    text = text.strip()
    if not text.endswith(')') and not text.endswith('）'):
        return text
    
    # 从末尾向前找到匹配的左括号
    close_char = text[-1]
    open_char = '(' if close_char == ')' else '（'
    
    depth = 0
    for i in range(len(text) - 1, -1, -1):
        if text[i] == close_char:
            depth += 1
        elif text[i] == open_char:
            depth -= 1
            if depth == 0:
                # 检查这个左括号前面是否有空格（表示是独立的括号段）
                prefix = text[:i].rstrip()
                return prefix
    
    return text


def extract_chinese_part(full_desc):
    """
    从完整描述中提取中文部分。
    格式：[英文描述] [中文类目] [中文商品名] [储存类型] [国家] ([重复中文部分])
    
    策略：通过匹配已知类目来定位中文部分的起始位置，
    避免英文部分中偶尔出现的中文字符（如"度"）造成误判。
    """
    if not full_desc:
        return ''
    
    # 先转换繁体为简体
    text = cc.convert(full_desc)
    
    # 移除末尾的括号重复部分（可能需要移除多次，如末尾有(集团合同)等注释）
    text = remove_trailing_parenthetical(text)
    text = remove_trailing_parenthetical(text)
    
    # 策略1：通过已知类目定位中文部分起始位置
    # 在文本中搜索已知类目，找到最靠前的匹配
    best_pos = -1
    best_cat = ''
    for cat in CATEGORIES_SORTED:
        # 使用词边界匹配：类目前面应该是空格或字符串开头
        pattern = r'(?:^|\s)' + re.escape(cat) + r'(?:\s|$)'
        m = re.search(pattern, text)
        if m:
            pos = m.start()
            # 跳过字符串开头的匹配（可能是英文部分）
            if best_pos == -1 or pos < best_pos:
                best_pos = pos
                best_cat = cat
    
    if best_pos >= 0 and best_cat:
        # 从类目位置开始提取
        chinese_part = text[best_pos:].strip()
        return chinese_part
    
    # 策略2：如果没找到已知类目，回退到找第一个中文字符
    match = re.search(r'[\u4e00-\u9fff]', text)
    if not match:
        return ''
    
    chinese_part = text[match.start():].strip()
    return chinese_part


def remove_category(text):
    """移除开头的类目信息"""
    text = text.strip()
    
    # 尝试匹配最长的类目
    for cat in CATEGORIES_SORTED:
        if text.startswith(cat):
            remaining = text[len(cat):].strip()
            if remaining:
                # 检查是否还有子类目
                for sub in SUB_CATEGORIES:
                    if remaining.startswith(sub):
                        remaining = remaining[len(sub):].strip()
                        break
                # 关键判断：如果移除类目后，剩余内容以括号（品牌）开头，
                # 说明类目本身就是商品名称（如"日本清酒 (菊正宗) 1800ML"），
                # 此时不应移除类目
                if remaining.startswith('(') or remaining.startswith('（'):
                    return text
                # 关键判断：如果类目是"黄油"且剩余部分以修饰词（如"无盐"）开头，
                # 说明类目是产品名核心部分（如"黄油 无盐 (总统)"即"无盐黄油"），应保留
                if cat == '黄油' and remaining.split(None, 1)[0] in ('无盐', '咸味', '淡味', '原味'):
                    return text
                # 关键判断：如果移除类目后，剩余部分以单个汉字开头（如"方""片"等形态词），
                # 说明类目是商品名不可分割的一部分（如"火腿 方"即"火腿方"），应保留类目
                _next_token = remaining.split(None, 1)[0]
                if re.match(r'^[\u4e00-\u9fff]$', _next_token):
                    return text
                return remaining
            else:
                return text  # 如果移除类目后为空，保留原文
    
    # 如果没有匹配到已知类目，尝试移除第一个空格前的内容（第一个token）
    parts = text.split(None, 1)
    if len(parts) > 1:
        # 检查第一个token是否看起来像类目（纯中文，较短）
        first_token = parts[0]
        if re.match(r'^[\u4e00-\u9fff]+$', first_token) and len(first_token) <= 6:
            # 同样检查：如果剩余部分以括号开头，保留原文
            if parts[1].startswith('(') or parts[1].startswith('（'):
                return text
            return parts[1]
    
    return text


def remove_suffix(text):
    """移除末尾的储存类型和国家信息"""
    text = text.strip()
    
    # 反复尝试移除末尾的国家名
    for country in sorted(COUNTRIES, key=len, reverse=True):
        if text.endswith(country):
            text = text[:-len(country)].strip()
            break
    
    # 移除末尾的储存类型
    for storage in STORAGE_TYPES:
        if text.endswith(storage):
            text = text[:-len(storage)].strip()
            break
    
    # 再次检查国家（有些格式是 储存类型 在国家前面，有些在后面）
    for country in sorted(COUNTRIES, key=len, reverse=True):
        if text.endswith(country):
            text = text[:-len(country)].strip()
            break
    
    return text


def remove_per_phrases(text):
    """移除 PER PKT, PER BTL, PER BAG, PER PC, PER BOX 等包装描述"""
    # 移除 "PER XXX" 模式
    text = re.sub(r'\s+PER\s+\w+', '', text, flags=re.IGNORECASE)
    # 移除 "PER PC," 带逗号的情况
    text = re.sub(r'\s+PER\s+\w+,', '', text, flags=re.IGNORECASE)
    # 移除 "FZ" 标记（急冻标记）
    text = re.sub(r'\s+FZ\b', '', text)
    # 移除 "CH" 标记（冰鲜标记）
    text = re.sub(r'\s+CH\b', '', text)
    # 移除 "DR" 标记（干货标记）
    text = re.sub(r'\s+DR\b', '', text)
    # 移除 "BTL" 前缀（如 BTL430ML -> 430ML）
    text = re.sub(r'\bBTL(?=\d)', '', text)
    # 移除 "CHN" 等英文国家代码残留（2-3个大写字母）
    text = re.sub(r'\s+[A-Z]{2,3}\b', '', text)
    # 移除独立的数字产品编码（4-5位数字，不是尺寸的一部分）
    text = re.sub(r'\s+\d{4,5}\b', '', text)
    # 移除 "员工厨" 等备注
    text = re.sub(r'\s*员工厨', '', text)
    # 移除末尾的逗号和不完整的数量信息（如 "7GM, 288"）
    text = re.sub(r',\s*\d+\s*$', '', text)
    return text.strip()


def add_brand_suffix(text):
    """在括号内的品牌名后添加"牌"字"""
    # 处理半角括号: (品牌) -> (品牌牌)
    text = re.sub(r'\(([^)]+?)\)', lambda m: f'({m.group(1)}牌)' if not m.group(1).endswith('牌') else m.group(0), text)
    # 处理全角括号: （品牌） -> （品牌牌）
    text = re.sub(r'（([^）]+?)）', lambda m: f'（{m.group(1)}牌）' if not m.group(1).endswith('牌') else m.group(0), text)
    return text


def clean_product_name(text):
    """清理商品名称中的多余空格等"""
    # 移除多余空格
    text = re.sub(r'\s{2,}', ' ', text)
    # 移除首尾空格
    text = text.strip()
    return text


def process_description(desc):
    """
    处理单条商品描述，提取中文商品名称。
    
    示例：
    "SNACKS MAGIC CHILI (HUANG FEI HONG) 308GM DR CHN 小食 香脆椒 (黄飞红) 308GM 干货 中国 (小食 香脆椒 (黄飞红) 308GM 干货 中国)"
    -> "香脆椒 (黄飞红) 308GM"
    """
    if not desc:
        return ''
    
    # Step 1: 提取中文部分（含繁转简、去尾部括号）
    chinese_part = extract_chinese_part(desc)
    if not chinese_part:
        return ''
    
    # Step 2: 移除末尾的储存类型和国家
    product = remove_suffix(chinese_part)
    
    # Step 3: 移除开头的类目
    product = remove_category(product)
    
    # Step 4: 移除 PER XXX 等包装描述
    product = remove_per_phrases(product)
    
    # Step 5: 括号内品牌名加"牌"字
    product = add_brand_suffix(product)
    
    # Step 6: 清理
    product = clean_product_name(product)
    
    return product


def process_excel_file(input_path, output_path):
    """处理Excel文件，生成客户商品名称表"""
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    
    # 创建输出工作簿
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = 'Sheet1'
    
    # 写入表头
    out_ws.cell(1, 1, '序号')
    out_ws.cell(1, 2, '原始订单名称')
    out_ws.cell(1, 3, '客户商品名')
    
    row_num = 0
    skipped = 0
    
    for r in range(2, ws.max_row + 1):
        desc = ws.cell(r, 1).value
        if not desc:
            continue
        
        original_name = str(desc).strip()
        product_name = process_description(original_name)
        
        if product_name:
            row_num += 1
            out_ws.cell(row_num + 1, 1, row_num)
            out_ws.cell(row_num + 1, 2, original_name)
            out_ws.cell(row_num + 1, 3, product_name)
        else:
            skipped += 1
    
    out_wb.save(output_path)
    print(f'处理完成: {input_path}')
    print(f'  成功提取: {row_num} 条')
    print(f'  跳过（无中文）: {skipped} 条')
    print(f'  输出文件: {output_path}')
    return row_num


# ============ 测试用例 ============
def test():
    """测试用户提供的案例"""
    test_cases = [
        (
            "HAM PORK COOKED (YU RUN) 2KG PER PC CANTEEN CH CHN 火腿 方 (雨润) 2KG PER PC 员工厨 冰鲜 中国 (火腿 方 (雨润) 2KG PER PC 员工厨 冰鲜 中国)",
            "火腿 方 (雨润牌) 2KG"
        ),
        (
            "SNACKS MAGIC CHILI (HUANG FEI HONG) 308GM DR CHN 小食 香脆椒 (黄飞红) 308GM 干货 中国 (小食 香脆椒 (黄飞红) 308GM 干货 中国)",
            "香脆椒 (黄飞红牌) 308GM"
        ),
        (
            "SNACKS SUNFLOWER SEED (QIA QIA) 178GM PER PKT DR CHN 小食 原香瓜子 (洽洽) 178GM PER PKT 干货 中国 (小食 原香瓜子 (洽洽) 178GM PER PKT 干货 中国)",
            "原香瓜子 (洽洽牌) 178GM"
        ),
        (
            "SAUCE SOUR AND HOT SOY (KNORR) 468ML DR CHN 酱料 酸辣鲜露 (家乐) 468ML 干货 中国 (酱料 酸辣鲜露 (家乐) 468ML 干货 中国)",
            "酸辣鲜露 (家乐牌) 468ML"
        ),
        (
            "FRUIT DRIED LEMON SLICED DR CHN 干果 干柠檬片 干货 中国 (干果 干柠檬片 干货 中国)",
            "干柠檬片"
        ),
    ]
    
    print("=" * 60)
    print("测试用例验证：")
    print("=" * 60)
    all_pass = True
    for desc, expected in test_cases:
        result = process_description(desc)
        status = "✓" if result == expected else "✗"
        if result != expected:
            all_pass = False
        print(f"\n{status} 输入: {desc[:60]}...")
        print(f"  期望: {expected}")
        print(f"  实际: {result}")
    
    print(f"\n{'=' * 60}")
    print(f"测试结果: {'全部通过' if all_pass else '存在失败'}")
    print(f"{'=' * 60}")
    return all_pass


if __name__ == '__main__':
    import os
    import argparse

    # 先运行测试
    test()
    print("\n")

    # 解析命令行参数
    parser = argparse.ArgumentParser(description='处理香格里拉报价单，提取标准中文商品名称')
    parser.add_argument('input_files', nargs='*', help='输入文件路径（不指定则使用默认的两个报价单）')
    parser.add_argument('-o', '--output', help='输出目录（默认与输入文件同目录）')
    args = parser.parse_args()

    if args.input_files:
        # 使用指定的输入文件
        files_to_process = []
        for f in args.input_files:
            if not os.path.exists(f):
                print(f'警告: 文件不存在，跳过 - {f}')
                continue
            dirname = os.path.dirname(f)
            basename = os.path.splitext(os.path.basename(f))[0]
            out_dir = args.output or dirname
            out_name = os.path.join(out_dir, basename + '_商品名称表.xlsx')
            files_to_process.append((f, out_name))
    else:
        # 默认处理两个报价单文件
        base_dir = os.path.dirname(os.path.abspath(__file__))
        files_to_process = [
            (os.path.join(base_dir, '报价--杭州老香749.xlsx'), os.path.join(base_dir, '杭州老香749_商品名称表.xlsx')),
            (os.path.join(base_dir, '报价--杭州老香967.xlsx'), os.path.join(base_dir, '杭州老香967_商品名称表.xlsx')),
        ]

    for input_file, output_file in files_to_process:
        process_excel_file(input_file, output_file)
        print()
